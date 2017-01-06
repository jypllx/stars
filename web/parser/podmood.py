import sys
from config import BaseConfig
import os.path
from openpyxl import Workbook, load_workbook

"""
Manages the mapping betweend the iTunes categories and the Tootal moods.
Uses an 2 column xls file :
 * 1st column : iTunes categories
 * 2nd column : Tootak Moods
"""

class PodMood:

    def __init__(self):
        self.file = BaseConfig.MOOD_FILE

        # if not os.path.exists(self.file) :
        #     raise Exception('No file for %s' % self.file)

        wb=load_workbook(self.file)
        ws=wb.active

        self.moods=[]
        self.mapping={}

        line=2
        iCat=ws.cell(row=line, column=2).value

        while iCat is not None:
            mood=ws.cell(row=line, column=1).value

            iCat=iCat.strip()
            mood=mood.strip()

            if mood is not None:
                self.moods.append(mood)
                self.mapping[iCat]=mood
            else:
                raise Exeption()
            line+=1
            iCat=ws.cell(row=line, column=2).value

        self.moods = list(set(self.moods))        

    def get_mood(self, itunes_category):
        try :
            return self.mapping[itunes_category]
        except Exception as e:
            return None

if __name__ == "__main__":
    p = PodMood('./moods.xlsx')